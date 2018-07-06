import openpyxl as ox
import options as ops

wb = ox.load_workbook(ops.xls_file)
# ws = wb.get_sheet_by_name("Sheet3")


class file:
    #하나의 관측에 대해 여러개의 파일이 있음 // 이런식으로 H1 Acc. File Name	H2 Acc. File Name	Vertical Acc. File Name
    #그래서 변경하고자 하는 이름을 c_name 거기에 해당하는 파일리스트(3개를 list에 집어넣음
    def __init__(self, c_name, f_list):
        self.c_name = c_name
        self.f_list = f_list

    def add_list(self, f_name):
        self.f_list.append(f_name)


class data:
    def __init__(self, num, region, city, year, magnitude):
        self.num = num
        self.region = region
        self.city = city
        self.year = year
        self.magnitude = magnitude
        #self.file_order = file_order #파일순서는 여기서 들어갈게 아니라 생각해서 BAAM


def get_data_list():
    data_list = list()

    # print(get_data_by_key(1))  #1번은 attribute name이 나옴 2번부터 진짜 data임
    i = 2
    while True:
        # xls_handler로 부터 reference.xlsx의 sheet1의 숫자를 기준으로 데이터 받음
        tmp_list = get_data_by_key(i)
        if tmp_list[0] == None:
            break
        else:
            # data 리스트에 넣는다
            # 순서, 지역, 도시, 년도, 규모*100(소수점이라), 해당 지역 파일순서 순
            # data_list.append(data(i - 1, ops.region, tmp_list[0], tmp_list[1], int(tmp_list[2] * 100), (i - 2) % 3 + 1))
            data_list.append(data(i - 1, ops.region, tmp_list[0], tmp_list[1], int(tmp_list[2] * 100)))
            i = i + 1

    return data_list


'''
def print_data_list(data_list):
    for d in data_list:
        print(d.num, d.region, d.city, d.year, d.magnitude)
'''


def get_city_dic():
    #파일에 입력된 도시이름을 파일형식에 맞게 간단하게 만드는 함수
    city_dic = {}
    ws = wb["Sheet3"]
    i = 2
    while True:
        key = ws['A' + str(i)].value
        value = ws['B' + str(i)].value
        if key == None:
            break
        else:
            city_dic[key] = value
            i = i + 1

    return city_dic


def get_data_by_key(key):
    # Event(city):J , Year:K, Mag:M
    ws = wb["Sheet1"]
    key = str(key)
    city_dic = get_city_dic()
    city_name = city_dic.get(ws['J' + key].value)
    return [city_name, ws['K' + key].value, ws['M' + key].value]


def make_file_list():
    # T : file1, U : file2, V : file3
    ws = wb["Sheet1"]
    c_list = get_file_changed_file_name_list()
    f_list = list()

    i = 2
    while True:
        if ws['A' + str(i)].value == None:
            break
        else:
            f_list.append(file(c_list[i-2],[ws['T'+str(i)].value, ws['U'+str(i)].value, ws['V'+str(i)].value]))
            i = i + 1

    return f_list

#######################################################################################


def get_file_changed_file_name_list():
    result = list()
    l = get_data_list()
    sep = ops.separator
    for lst in l:
        result.append(
            str(lst.num) + sep + lst.region + sep + lst.city + sep +
            str(lst.year) + sep + str(lst.magnitude))

    return result


def print_lsit():
    for l in make_file_list():
        print(l.c_name, l.f_list)


wb.close()

# 1 : num
# 2 : region
# 3 : city
# 4 : year
# 5 : magnitude
# 6 : file order
